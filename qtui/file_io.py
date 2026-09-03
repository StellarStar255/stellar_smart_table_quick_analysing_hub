# -*- coding: utf-8 -*-
"""
文件读写层 - 从 mixins/file_mixin.py 提取的纯数据逻辑（pandas + openpyxl）。

不涉及任何 GUI；主窗口通过后台线程调用这些函数。
"""

import datetime
import json
import os
import re
import stat
import tempfile

import numpy as np
import pandas as pd

from qtui.i18n import tr

RECENT_FILES_PATH = os.path.expanduser("~/.smart_table_hub/recent_files.json")
MAX_RECENT_FILES = 10

# 尝试顺序：utf-8-sig 兼容有/无 BOM 的 utf-8；gb18030 是 gbk 的超集但几乎
# 对任何字节对都不报错，必须排在 big5 之后；latin1 对任何字节都不报错，
# 只能放最后兜底。
CSV_ENCODINGS = ("utf-8-sig", "gbk", "big5", "gb18030", "utf-16", "latin1")

# 新建 CSV 的默认编码：带 BOM 的 utf-8，Windows 上的 Excel 才能正确识别中文
DEFAULT_CSV_ENCODING = "utf-8-sig"

# Excel 对 sheet 名的限制
SHEET_NAME_MAX_LEN = 31
_SHEET_NAME_BAD_CHARS = set('\\/?*[]:')


# ---------- 加载 ----------

def load_workbook_lazy(file_path):
    """打开 Excel 文件，返回 (ExcelFile, sheet_names)。不加载任何 sheet 数据。

    优先用 calamine 引擎（Rust 实现，大文件比 openpyxl 快数倍），
    不可用或打开失败时回退到 pandas 默认引擎。
    """
    try:
        excel_file = pd.ExcelFile(file_path, engine="calamine")
    except Exception:
        excel_file = pd.ExcelFile(file_path)
    return excel_file, list(excel_file.sheet_names)


def read_sheet(excel_file: pd.ExcelFile, sheet_name: str) -> pd.DataFrame:
    """读取单个 sheet。

    先按单元格原始类型读入（dtype=object），再让 pandas 只推断整列同类型的
    列：数字列仍是 float64/int64、日期列仍是 datetime64，而 Excel 里以文本
    格式存放的 "007"（工号/邮编/编码）保持字符串，不会被静默改成 7。
    重复表头 pandas 已自动改名（A、A.1 …）。
    """
    df = excel_file.parse(sheet_name, dtype=object).infer_objects()
    df.columns = [str(c) for c in df.columns]
    return df


def _read_full_ragged_csv(file_path, enc, sep):
    """整文件按最大行宽原样读入：不丢弃任何行（含元数据前言与空行），
    列名用位置字母占位。

    世界银行等导出的 CSV 前几行比数据区窄，pandas 按首行推断列数会
    解析失败。这里只负责"完整载入"；是否把某行提升为表头由用户在
    界面上自行决定（绝不静默删行）。
    """
    import csv as _csv
    width = 0
    with open(file_path, "r", encoding=enc, newline="") as f:
        for row in _csv.reader(f, delimiter=sep):
            width = max(width, len(row))
    if width == 0:
        return None
    from core.formula_engine import FormulaEngine
    names = [FormulaEngine.col_index_to_letter(i) for i in range(width)]
    kwargs = dict(encoding=enc, sep=sep, header=None, names=names,
                  skip_blank_lines=False)
    df = pd.read_csv(file_path, **kwargs)
    return _preserve_leading_zeros(df, file_path, kwargs)


def _sniff_bom(file_path):
    """按文件头 BOM 判断编码；无 BOM 返回 None。"""
    try:
        with open(file_path, "rb") as f:
            head = f.read(4)
    except OSError:
        return None
    if head.startswith(b"\xef\xbb\xbf"):
        return "utf-8-sig"
    if head.startswith((b"\xff\xfe", b"\xfe\xff")):
        return "utf-16"
    return None


def _csv_may_have_leading_zeros(file_path, sep):
    """字节级预扫：文件里是否出现"分隔符/引号/换行 + 0 + 数字"的字段开头。

    绝大多数 CSV 没有前导零字段，预扫（memchr 速度）比二次解析便宜得多；
    只有命中时才对数值列做逐列核对。
    """
    sep_b = sep.encode("ascii", "replace") if sep else b","
    pat = re.compile(rb'[' + re.escape(sep_b) + rb'"\n]-?0[0-9]')
    try:
        with open(file_path, "rb") as f:
            first = f.read(3)
            if first[:1] == b"0" and first[1:2].isdigit():
                return True
            if first[:1] == b"-" and first[1:2] == b"0" and first[2:3].isdigit():
                return True
            f.seek(0)
            tail = b""
            while True:
                chunk = f.read(1 << 22)
                if not chunk:
                    return False
                buf = tail + chunk
                if pat.search(buf):
                    return True
                tail = buf[-4:]
    except OSError:
        return True


def _preserve_leading_zeros(df, file_path, read_kwargs):
    """把被 pandas 推断成数字、但原文带前导零的列（邮编/工号/编码）还原为文本。

    做法：只对"整数值"列（int 列，或非空值全为整数的 float 列）用 dtype=str
    重读，比较原文长度与数字位数——原文更长即存在前导零（或 + 号），整列
    保持原文。不改动其它列，pandas 的快速推断路径保持不变。
    """
    if df is None or len(df) == 0 or len(df.columns) == 0:
        return df
    enc = read_kwargs.get("encoding") or ""
    if not enc.lower().startswith("utf-16"):
        if not _csv_may_have_leading_zeros(file_path, read_kwargs.get("sep") or ","):
            return df
    cand = []
    for i in range(len(df.columns)):
        s = df.iloc[:, i]
        if pd.api.types.is_bool_dtype(s):
            continue
        if pd.api.types.is_integer_dtype(s):
            cand.append(i)
        elif pd.api.types.is_float_dtype(s):
            v = s.to_numpy()
            nn = v[~np.isnan(v)]
            if len(nn) and np.all(nn == np.floor(nn)):
                cand.append(i)
    if not cand:
        return df
    try:
        raw = pd.read_csv(file_path, usecols=cand, dtype=str, **read_kwargs)
    except Exception:
        return df
    if len(raw) != len(df):
        return df
    for pos, i in enumerate(cand):
        rs = raw.iloc[:, pos]
        mask = rs.notna().to_numpy()
        if not mask.any():
            continue
        texts = rs.to_numpy()[mask].astype(str)
        lens = np.char.str_len(np.char.strip(texts))
        v = df.iloc[:, i].to_numpy()[mask].astype(np.float64)
        av = np.abs(v)
        digits = np.where(av < 1, 1, np.floor(np.log10(np.maximum(av, 1)) + 1e-9) + 1)
        digits = digits + (v < 0)
        if (lens > digits).any():
            df.isetitem(i, rs.to_numpy())
    return df


def read_csv_any_encoding(file_path, delimiter=None) -> pd.DataFrame:
    """按 BOM → utf-8 → gbk → big5 → gb18030 → utf-16 → latin1 顺序尝试读取 CSV/TSV。

    列数不一致（前言元数据行比数据行窄）导致解析失败时，改为整文件
    原样载入（首行也作为数据、列名用位置字母），由用户决定表头。
    空文件返回空 DataFrame。检测到的编码记录在 df.attrs["source_encoding"]，
    保存时可按原编码写回。
    """
    last_err = None
    sep = delimiter
    if sep is None:
        sep = "\t" if file_path.lower().endswith((".tsv", ".txt")) else ","
    bom_enc = _sniff_bom(file_path)
    encodings = (bom_enc,) if bom_enc else CSV_ENCODINGS
    for enc in encodings:
        kwargs = dict(encoding=enc, sep=sep)
        try:
            df = pd.read_csv(file_path, **kwargs)
            df = _preserve_leading_zeros(df, file_path, kwargs)
        except UnicodeDecodeError as e:
            last_err = e
            continue
        except UnicodeError as e:  # utf-16 缺 BOM 等
            last_err = e
            continue
        except pd.errors.EmptyDataError:
            df = pd.DataFrame()
        except pd.errors.ParserError as e:
            last_err = e
            try:
                df = _read_full_ragged_csv(file_path, enc, sep)
            except (UnicodeDecodeError, UnicodeError, pd.errors.ParserError) as e2:
                last_err = e2
                continue
            if df is None:
                df = pd.DataFrame()
        df.attrs["source_encoding"] = _source_encoding_name(file_path, enc)
        return df
    raise last_err


def _source_encoding_name(file_path, enc):
    """utf-8-sig 能解码无 BOM 的 utf-8；记录时区分有无 BOM，保存才能原样写回。"""
    if enc == "utf-8-sig" and _sniff_bom(file_path) != "utf-8-sig":
        return "utf-8"
    return enc


def _xlsx_has_formulas(file_path):
    """直接扫 xlsx 压缩包里 worksheet XML 的 <f> 公式标签。

    C 速度的字节扫描，远快于 openpyxl 逐格解析；绝大多数纯数据文件
    没有公式，可借此完全跳过公式扫描。
    """
    import zipfile
    try:
        with zipfile.ZipFile(file_path) as z:
            for name in z.namelist():
                if not (name.startswith("xl/worksheets/") and name.endswith(".xml")):
                    continue
                with z.open(name) as f:
                    tail = b""
                    while True:
                        chunk = f.read(1 << 20)
                        if not chunk:
                            break
                        buf = tail + chunk
                        if b"<f>" in buf or b"<f " in buf:
                            return True
                        tail = buf[-3:]  # 防止标签跨块边界
    except Exception:
        return True  # 无法判断时保守走完整扫描
    return False


def read_sheet_formulas(file_path, sheet_name) -> dict:
    """用 openpyxl 扫描 sheet 中的公式，返回 {(row, col): "=..."}。

    pandas 读到的是公式的缓存计算值，公式文本必须用 data_only=False 另读。
    仅支持 .xlsx；读取失败返回空 dict。
    """
    if not str(file_path).lower().endswith(".xlsx"):
        return {}
    if not _xlsx_has_formulas(file_path):
        return {}
    formulas = {}
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True, data_only=False)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return {}
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2):  # 第 1 行是表头
            for cell in row:
                v = cell.value
                if isinstance(v, str) and len(v) > 1 and v.startswith("="):
                    formulas[(cell.row - 2, cell.column - 1)] = v
        wb.close()
    except Exception as e:
        print(f"读取公式失败: {e}")
    return formulas


def _xlsx_has_custom_fills(file_path) -> bool:
    """styles.xml 里默认只有 none+gray125 两个 fill，更多说明存在自定义填充。

    字节级预检，避免为绝大多数无背景色的文件做整表样式扫描。
    """
    import zipfile
    try:
        with zipfile.ZipFile(file_path) as z:
            with z.open("xl/styles.xml") as f:
                return f.read().count(b"<fill>") > 2
    except (OSError, KeyError, zipfile.BadZipFile, TypeError,
            ValueError, AttributeError):
        return False


def read_sheet_colors(file_path, sheet_name) -> dict:
    """读取 sheet 的单元格背景色，返回 {(数据行, 列): '#rrggbb'}。

    行号 -1 表示表头行（Excel 第 1 行）。仅支持 .xlsx；
    无自定义填充或读取失败返回空 dict。
    """
    if not str(file_path).lower().endswith(".xlsx"):
        return {}
    if not _xlsx_has_custom_fills(file_path):
        return {}
    colors = {}
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return {}
        for row in wb[sheet_name].iter_rows():
            for cell in row:
                fill = cell.fill
                if fill is None or fill.fill_type != "solid":
                    continue
                rgb = getattr(fill.fgColor, "rgb", None)
                if not isinstance(rgb, str) or rgb == "00000000":
                    continue
                colors[(cell.row - 2, cell.column - 1)] = "#" + rgb[-6:].lower()
        wb.close()
    except Exception as e:
        print(f"读取背景色失败: {e}")
    return colors


def _dedupe_headers(headers):
    seen = {}
    result = []
    for h in headers:
        if h in seen:
            seen[h] += 1
            result.append(f"{h}.{seen[h]}")
        else:
            seen[h] = 0
            result.append(h)
    return result


# ---------- 保存 ----------

# 坐标版本标记（写入 xlsx 文档属性 keywords）：带标记 = 本应用新坐标系
# （第 1 行表头、公式坐标与 Excel 一致）保存，加载时公式结果可放心写入；
# 无标记 = Excel/旧版应用来源，加载时错误结果不覆盖文件缓存值
COORD_MARKER = "SmartTableHub-coord-v2"


def xlsx_has_coord_marker(file_path) -> bool:
    """检查 xlsx 是否带本应用的坐标版本标记（直接扫 docProps/core.xml）。"""
    import zipfile
    try:
        with zipfile.ZipFile(file_path) as z:
            with z.open("docProps/core.xml") as f:
                return COORD_MARKER.encode() in f.read()
    except (OSError, KeyError, zipfile.BadZipFile,
            TypeError, ValueError, AttributeError):
        return False


def check_sheet_name(name, existing_names=()):
    """校验 sheet 名是否符合 Excel 规则；合法返回 None，否则返回错误说明。

    规则：非空、不超过 31 个字符、不含 \\ / ? * [ ] :、
    与 existing_names 不重名（Excel 不区分大小写）。
    """
    if name is None or not str(name).strip():
        return tr("Sheet 名不能为空")
    name = str(name)
    if len(name) > SHEET_NAME_MAX_LEN:
        return tr("Sheet 名不能超过 {} 个字符").format(SHEET_NAME_MAX_LEN)
    bad = [c for c in name if c in _SHEET_NAME_BAD_CHARS]
    if bad:
        return tr("Sheet 名不能包含字符: {}").format(" ".join(sorted(set(bad))))
    lowered = name.lower()
    for other in existing_names:
        if str(other).lower() == lowered:
            return tr("已存在同名 Sheet: {}").format(other)
    return None


def _replace_file(tmp_path, file_path):
    """用临时文件原子替换目标；保留目标原有权限位（新文件用 0644）。"""
    try:
        mode = stat.S_IMODE(os.stat(file_path).st_mode)
    except OSError:
        mode = 0o644
    try:
        os.chmod(tmp_path, mode)
    except OSError:
        pass
    # os.replace 在 POSIX 与 Windows 上都是原子替换；失败时直接报错，
    # 绝不退化成"先清空目标再拷贝"的非原子写法
    os.replace(tmp_path, file_path)


def save_workbook(file_path, sheets: dict, sheet_order=None, formulas=None,
                  progress_cb=None, cell_colors=None):
    """把 {sheet名: DataFrame} 全量写入 xlsx。

    formulas: 可选 {sheet名: {(row, col): "=..."}}，公式覆盖写入对应单元格，
    Excel 打开时仍是可计算的公式（df 中已存计算结果，作为兜底值先写入）。
    cell_colors: 可选 {sheet名: {(数据行, 列): '#rrggbb'}}，行 -1 为表头行，
    写成真实的单元格填充（Excel 中同样可见）。
    progress_cb: 可选 (sheet名, 序号从1起, 总数) -> None，逐 sheet 汇报进度。
    先写临时文件再原子替换，避免写一半损坏原文件（与旧版后台保存策略一致）。
    sheet 名不合法（过长/非法字符/重名）时抛 ValueError，而不是静默截断
    导致两个 sheet 互相覆盖。
    """
    from openpyxl.styles import PatternFill
    order = sheet_order or list(sheets.keys())
    order = [n for n in order if n in sheets]
    for i, name in enumerate(order):
        err = check_sheet_name(name, order[:i])
        if err:
            raise ValueError(tr("Sheet 名 \"{}\" 不合法：{}").format(name, err))
    formulas = formulas or {}
    cell_colors = cell_colors or {}
    fd, tmp_path = tempfile.mkstemp(suffix=".xlsx", dir=os.path.dirname(file_path) or ".")
    os.close(fd)
    try:
        fill_cache = {}
        with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
            for i, name in enumerate(order):
                if progress_cb:
                    progress_cb(name, i + 1, len(order))
                sheets[name].to_excel(writer, sheet_name=name, index=False)
                for (row, col), formula in formulas.get(name, {}).items():
                    # +2: 跳过表头行且 openpyxl 从 1 开始计数
                    writer.sheets[name].cell(row=row + 2, column=col + 1,
                                             value=formula)
                for (row, col), color in cell_colors.get(name, {}).items():
                    argb = "FF" + str(color).lstrip("#").upper()
                    fill = fill_cache.get(argb)
                    if fill is None:
                        fill = PatternFill(start_color=argb, end_color=argb,
                                           fill_type="solid")
                        fill_cache[argb] = fill
                    excel_row = row + 2 if row >= 0 else 1   # -1 = 表头行
                    writer.sheets[name].cell(
                        row=excel_row, column=col + 1).fill = fill
            # 坐标版本标记：本应用保存的文件加载时公式结果可放心写入
            writer.book.properties.keywords = COORD_MARKER
        _replace_file(tmp_path, file_path)
    except Exception:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
        raise


# openpyxl 读不回来、就地保存会丢的部件（图表/图片/条件格式/数据验证等
# 它是能保留的，下面这些不行）。扫到就在保存前提示用户。
_LOSSY_PARTS = (
    ("xl/slicers/", "切片器"),
    ("xl/slicerCaches/", "切片器"),
    ("xl/timelines/", "日程表"),
    ("xl/ctrlProps/", "窗体控件"),
    ("xl/activeX/", "ActiveX 控件"),
    ("xl/threadedComments/", "新版批注（讨论）"),
    ("xl/richData/", "单元格内图片/富数据"),
    ("customXml/", "自定义 XML"),
    ("vbaProject.bin", "宏（VBA）"),
)


def xlsx_lossy_parts(file_path):
    """扫描 xlsx，返回就地保存时会丢失的功能名（去重、保持顺序）。"""
    import zipfile
    found = []
    try:
        with zipfile.ZipFile(file_path) as z:
            names = z.namelist()
    except Exception:
        return []
    for part, label in _LOSSY_PARTS:
        if label in found:
            continue
        if any(part in n for n in names):
            found.append(label)
    return found


def _xl_value(v):
    """numpy/pandas 标量 -> openpyxl 能写的 Python 类型；缺失值 -> None。"""
    if v is None or (isinstance(v, float) and v != v):
        return None
    if isinstance(v, (str, bool, int, float)):
        return v
    try:
        if pd.isna(v):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(v, pd.Timestamp):
        return v.to_pydatetime()
    if isinstance(v, np.datetime64):
        return pd.Timestamp(v).to_pydatetime()
    if isinstance(v, np.generic):
        return v.item()
    return v if isinstance(v, (datetime.datetime, datetime.date,
                               datetime.time)) else str(v)


def _fill_for(color, cache):
    from openpyxl.styles import PatternFill
    argb = "FF" + str(color).lstrip("#").upper()
    fill = cache.get(argb)
    if fill is None:
        fill = PatternFill(start_color=argb, end_color=argb, fill_type="solid")
        cache[argb] = fill
    return fill


def _write_sheet_values(ws, df, clear_stale_fills):
    """把 DataFrame 整块覆盖到 sheet（第 1 行表头），并删掉多余的行列。

    表头行原本是公式的单元格不覆盖，否则用户的公式会被写死成静态文本；
    返回这样保留下来的表头公式个数。
    clear_stale_fills=True 时顺手清掉数据区里的旧底色（用户清除颜色才生效）。
    """
    kept = 0
    ncols = len(df.columns)
    nrows = len(df.index)
    for j, name in enumerate(df.columns):
        cell = ws.cell(row=1, column=j + 1)
        if isinstance(cell.value, str) and cell.value.startswith("="):
            kept += 1
            continue
        cell.value = str(name)
    for i, row in enumerate(df.itertuples(index=False, name=None)):
        for j, v in enumerate(row):
            # 不能用 ws.cell(..., value=...)：openpyxl 对 None 是跳过不写，
            # 旧值会留在格子里，用户删掉的内容就删不掉
            ws.cell(row=i + 2, column=j + 1).value = _xl_value(v)
    if clear_stale_fills:
        from openpyxl.styles import PatternFill
        none_fill = PatternFill()
        for row in ws.iter_rows(min_row=1, max_row=nrows + 1, max_col=ncols):
            for cell in row:
                if cell.fill is not None and cell.fill.fill_type is not None:
                    cell.fill = none_fill
    # 多余的行列真删掉（表变短/变窄时不留空壳）
    if ws.max_row > nrows + 1:
        ws.delete_rows(nrows + 2, ws.max_row - (nrows + 1))
    if ws.max_column > ncols:
        ws.delete_cols(ncols + 1, ws.max_column - ncols)
    return kept


def _apply_fills(ws, colors, cache):
    for (row, col), color in colors.items():
        excel_row = row + 2 if row >= 0 else 1   # -1 = 表头行
        ws.cell(row=excel_row, column=col + 1).fill = _fill_for(color, cache)


def patch_workbook(src_path, dest_path, sheets: dict, sheet_order,
                   formulas=None, cell_colors=None, progress_cb=None):
    """在原工作簿基础上就地更新数据后另存，保留 pandas 重建会丢掉的一切
    （透视表、条件格式、数据验证、数字格式、列宽、合并单元格、图表、图片…）。

    sheets: {sheet名: DataFrame}，只写这些 sheet；不在里面的 sheet 原样保留。
    sheet_order: 保存后完整的 sheet 顺序；不在其中的 sheet 视为用户删除。
    formulas / cell_colors: 与 save_workbook 同义，只对 sheets 里的 sheet 生效。
    返回 {"kept_header_formulas": n}。
    先写临时文件再原子替换，写一半失败不会损坏任何一个文件。
    """
    from openpyxl import load_workbook
    order = list(sheet_order)
    for i, name in enumerate(order):
        err = check_sheet_name(name, order[:i])
        if err:
            raise ValueError(tr("Sheet 名 \"{}\" 不合法：{}").format(name, err))
    formulas = formulas or {}
    cell_colors = cell_colors or {}
    # 原文件本来就没有自定义底色时，不必为"清除底色"去扫整个数据区
    clear_stale_fills = _xlsx_has_custom_fills(src_path)
    keep_vba = (str(src_path).lower().endswith((".xlsm", ".xltm"))
                and str(dest_path).lower().endswith((".xlsm", ".xltm")))
    wb = load_workbook(src_path, data_only=False, keep_vba=keep_vba)
    kept_header_formulas = 0
    fd, tmp_path = tempfile.mkstemp(suffix=os.path.splitext(dest_path)[1] or ".xlsx",
                                    dir=os.path.dirname(dest_path) or ".")
    os.close(fd)
    try:
        for name in list(wb.sheetnames):
            if name not in order:
                del wb[name]                      # 应用里删掉的 sheet
        fill_cache = {}
        for i, name in enumerate(order):
            if progress_cb:
                progress_cb(name, i + 1, len(order))
            df = sheets.get(name)
            if df is None:
                continue                          # 没改动过：整张原样保留
            ws = wb[name] if name in wb.sheetnames else wb.create_sheet(title=name)
            kept_header_formulas += _write_sheet_values(ws, df, clear_stale_fills)
            for (row, col), formula in formulas.get(name, {}).items():
                # +2: 跳过表头行且 openpyxl 从 1 开始计数
                ws.cell(row=row + 2, column=col + 1, value=formula)
            _apply_fills(ws, cell_colors.get(name, {}), fill_cache)
        wb._sheets = [wb[n] for n in order if n in wb.sheetnames]
        wb.properties.keywords = COORD_MARKER
        wb.save(tmp_path)
        wb.close()
        _replace_file(tmp_path, dest_path)
    except Exception:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
        raise
    return {"kept_header_formulas": kept_header_formulas}


def save_csv(file_path, df: pd.DataFrame, encoding=None):
    """写 CSV：先写同目录临时文件再原子替换，写一半失败不会损坏原文件。

    encoding 未指定时依次取：df.attrs["source_encoding"]（读取时记录的
    原文件编码，按原样写回）→ utf-8-sig（新文件；带 BOM，Windows Excel
    才能正确显示中文）。
    """
    enc = encoding or df.attrs.get("source_encoding") or DEFAULT_CSV_ENCODING
    fd, tmp_path = tempfile.mkstemp(suffix=".csv", dir=os.path.dirname(file_path) or ".")
    os.close(fd)
    try:
        df.to_csv(tmp_path, index=False, encoding=enc)
        _replace_file(tmp_path, file_path)
    except Exception:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
        raise


# ---------- 最近文件 ----------

def load_recent_files():
    """返回 (recent_files 列表, auto_save 布尔)。"""
    try:
        with open(RECENT_FILES_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)
        return data.get("recent_files", []), data.get("auto_save", False)
    except (OSError, json.JSONDecodeError):
        return [], False


def save_recent_files(recent_files, auto_save):
    os.makedirs(os.path.dirname(RECENT_FILES_PATH), exist_ok=True)
    with open(RECENT_FILES_PATH, "w", encoding="utf-8") as f:
        json.dump({"recent_files": recent_files, "auto_save": auto_save},
                  f, ensure_ascii=False, indent=2)


def add_recent_file(recent_files, path):
    path = os.path.abspath(path)
    files = [p for p in recent_files if p != path]
    files.insert(0, path)
    return files[:MAX_RECENT_FILES]
