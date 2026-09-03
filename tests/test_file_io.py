"""文件读取测试：带元数据前言的 CSV（世界银行式导出）——原样载入，不静默删行"""
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import pandas as pd

from qtui import file_io
from qtui.main_window import MainWindow

WORLD_BANK_CSV = (
    '"数据源","世界发展指标",\n'
    '\n'
    '"最后更新时间","2026-07-13",\n'
    '\n'
    '"Country Name","Country Code","Indicator Name","Indicator Code","1960","1961"\n'
    '"阿鲁巴","ABW","GDP","NY.GDP.MKTP.CD","",""\n'
    '"阿富汗","AFG","GDP","NY.GDP.MKTP.CD","1000","2000"\n'
    '"安哥拉","AGO","GDP","NY.GDP.MKTP.CD","3000","4000"\n'
)


class TestPreambleCsvFullLoad:
    def test_world_bank_csv_loads_everything(self, tmp_path):
        # 原样载入：元数据行、空行一行不丢，列名用位置字母，由用户决定表头
        path = str(tmp_path / 'wb.csv')
        with open(path, 'w', encoding='utf-8') as f:
            f.write(WORLD_BANK_CSV)
        df = file_io.read_csv_any_encoding(path)
        assert list(df.columns) == ['A', 'B', 'C', 'D', 'E', 'F']
        assert len(df) == 8                       # 2 元数据 + 2 空行 + 表头 + 3 数据
        assert df.iloc[0, 0] == '数据源'           # 前言仍在
        assert df.iloc[4, 0] == 'Country Name'    # 真表头作为数据行保留
        assert df.iloc[7, 0] == '安哥拉'

    def test_gbk_preamble_csv_loads(self, tmp_path):
        path = str(tmp_path / 'g.csv')
        with open(path, 'w', encoding='gbk') as f:
            f.write(WORLD_BANK_CSV)
        df = file_io.read_csv_any_encoding(path)
        assert len(df) == 8
        assert df.iloc[4, 0] == 'Country Name'

    def test_normal_csv_unaffected(self, tmp_path):
        path = str(tmp_path / 'n.csv')
        with open(path, 'w', encoding='utf-8') as f:
            f.write('A,B\n1,2\n3,4\n')
        df = file_io.read_csv_any_encoding(path)
        assert list(df.columns) == ['A', 'B']
        assert len(df) == 2


class TestHeaderCandidateDetection:
    def _full_loaded_df(self):
        return pd.DataFrame({
            'A': ['数据源', None, '最后更新时间', None, 'Country Name', '阿鲁巴', '阿富汗'],
            'B': ['世界发展指标', None, '2026-07-13', None, 'Country Code', 'ABW', 'AFG'],
            'C': [None, None, None, None, 'Indicator', 'GDP', 'GDP'],
            'D': [None, None, None, None, '1960', None, '1000'],
        })

    def test_detects_header_after_narrow_preamble(self):
        assert MainWindow._detect_header_candidate(self._full_loaded_df()) == 4

    def test_named_columns_not_flagged(self):
        df = pd.DataFrame({'姓名': ['a', 'b', 'c'], '金额': [1, 2, 3]})
        assert MainWindow._detect_header_candidate(df) == -1

    def test_letter_columns_without_preamble_not_flagged(self):
        df = pd.DataFrame({'A': ['x', 'y', 'z'], 'B': [1, 2, 3]})
        assert MainWindow._detect_header_candidate(df) == -1


class TestCellColorPersistence:
    """背景色写入 xlsx 真实填充并可读回（重启不丢、Excel 可见）"""

    def test_colors_round_trip(self, tmp_path):
        df = pd.DataFrame({'X': [1.0, 2.0], 'Y': [3.0, 4.0]})
        path = str(tmp_path / 'c.xlsx')
        colors = {(-1, 0): '#ff0000',   # 表头行
                  (0, 1): '#e3f2fd',
                  (1, 0): '#00ff00'}
        file_io.save_workbook(path, {'S': df}, ['S'],
                              cell_colors={'S': colors})
        assert file_io.read_sheet_colors(path, 'S') == colors

    def test_plain_file_fast_path_returns_empty(self, tmp_path):
        path = str(tmp_path / 'p.xlsx')
        file_io.save_workbook(path, {'S': pd.DataFrame({'X': [1.0]})}, ['S'])
        assert not file_io._xlsx_has_custom_fills(path)
        assert file_io.read_sheet_colors(path, 'S') == {}


class TestDetectPreambleEnd:
    def test_world_bank_style_rows(self):
        rows = [['数据源', '世界发展指标'],
                ['最后更新时间', '2026-07-13'],
                ['Country Name', 'Country Code', 'Indicator', '1960'],
                ['阿鲁巴', 'ABW', 'GDP', '100'],
                ['阿富汗', 'AFG', 'GDP', '200']]
        assert MainWindow._detect_preamble_end(rows) == 2

    def test_uniform_rows_no_preamble(self):
        rows = [['A', 'B'], ['1', '2'], ['3', '4']]
        assert MainWindow._detect_preamble_end(rows) == 0

    def test_too_few_rows(self):
        assert MainWindow._detect_preamble_end([['a'], ['b', 'c']]) == 0


class TestLeadingZerosPreserved:
    """文本格式的 "007"（工号/邮编）读入与保存都不能被静默改成数字 7"""

    def test_xlsx_text_cell_keeps_leading_zeros_and_round_trips(self, tmp_path):
        import openpyxl
        path = str(tmp_path / 'lz.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'S'
        ws.append(['ID', 'N'])
        for r, (sid, n) in enumerate([('007', 3), ('0123', 4.5)], start=2):
            c = ws.cell(row=r, column=1, value=sid)
            c.number_format = '@'
            ws.cell(row=r, column=2, value=n)
        wb.save(path)

        for engine in ('calamine', 'openpyxl'):
            xf = pd.ExcelFile(path, engine=engine)
            df = file_io.read_sheet(xf, 'S')
            assert df['ID'].tolist() == ['007', '0123'], engine
            assert df['N'].dtype.kind == 'f', engine   # 数字列仍是数字

        file_io.save_workbook(path, {'S': df}, ['S'])
        ws = openpyxl.load_workbook(path)['S']
        assert ws['A2'].value == '007'
        assert ws['B2'].value == 3

    def test_csv_leading_zero_column_stays_text(self, tmp_path):
        path = str(tmp_path / 'lz.csv')
        with open(path, 'w', encoding='utf-8') as f:
            f.write('zip,n,neg,f,mixed\n00123,1,-007,0.5,7\n00456,2,-8,1.5,08\n')
        df = file_io.read_csv_any_encoding(path)
        assert df['zip'].tolist() == ['00123', '00456']
        assert df['neg'].tolist() == ['-007', '-8']
        assert df['mixed'].tolist() == ['7', '08']
        assert df['n'].dtype.kind == 'i'      # 普通整数列不受影响
        assert df['f'].dtype.kind == 'f'      # "0.5" 不是前导零

    def test_csv_without_leading_zeros_infers_numbers(self, tmp_path):
        path = str(tmp_path / 'n.csv')
        with open(path, 'w', encoding='utf-8') as f:
            f.write('a,b\n10,0.5\n20,0\n')
        df = file_io.read_csv_any_encoding(path)
        assert df['a'].dtype.kind == 'i' and df['b'].dtype.kind == 'f'


class TestCsvEncoding:
    def test_utf16_bom_detected(self, tmp_path):
        path = str(tmp_path / 'u16.csv')
        with open(path, 'w', encoding='utf-16') as f:
            f.write('姓名,金额\n张三,10\n')
        df = file_io.read_csv_any_encoding(path)
        assert list(df.columns) == ['姓名', '金额']
        assert df.attrs['source_encoding'] == 'utf-16'

    def test_utf8_bom_detected_and_written_back(self, tmp_path):
        path = str(tmp_path / 'bom.csv')
        with open(path, 'w', encoding='utf-8-sig') as f:
            f.write('姓名,金额\n张三,10\n')
        df = file_io.read_csv_any_encoding(path)
        assert df.attrs['source_encoding'] == 'utf-8-sig'
        file_io.save_csv(path, df)
        with open(path, 'rb') as f:
            assert f.read(3) == b'\xef\xbb\xbf'
        assert list(file_io.read_csv_any_encoding(path).columns) == ['姓名', '金额']

    def test_plain_utf8_stays_without_bom(self, tmp_path):
        path = str(tmp_path / 'u8.csv')
        with open(path, 'w', encoding='utf-8') as f:
            f.write('姓名,金额\n张三,10\n')
        df = file_io.read_csv_any_encoding(path)
        assert df.attrs['source_encoding'] == 'utf-8'
        file_io.save_csv(path, df)
        with open(path, 'rb') as f:
            assert not f.read(3).startswith(b'\xef\xbb\xbf')

    def test_gbk_round_trip(self, tmp_path):
        path = str(tmp_path / 'g.csv')
        with open(path, 'w', encoding='gbk') as f:
            f.write('姓名,金额\n张三,10\n')
        df = file_io.read_csv_any_encoding(path)
        assert df.attrs['source_encoding'] == 'gbk'
        file_io.save_csv(path, df)
        with open(path, 'r', encoding='gbk') as f:
            assert f.readline().strip() == '姓名,金额'

    def test_big5_detected(self, tmp_path):
        path = str(tmp_path / 'b5.csv')
        with open(path, 'w', encoding='big5') as f:
            f.write('姓名,金額\n王五,10\n')
        df = file_io.read_csv_any_encoding(path)
        assert list(df.columns) == ['姓名', '金額']

    def test_new_csv_defaults_to_utf8_bom(self, tmp_path):
        path = str(tmp_path / 'new.csv')
        file_io.save_csv(path, pd.DataFrame({'a': ['中']}))
        with open(path, 'rb') as f:
            assert f.read(3) == b'\xef\xbb\xbf'


class TestEmptyCsv:
    def test_empty_file_returns_empty_frame(self, tmp_path):
        path = str(tmp_path / 'e.csv')
        open(path, 'w').close()
        df = file_io.read_csv_any_encoding(path)
        assert df.shape == (0, 0)

    def test_blank_lines_only(self, tmp_path):
        path = str(tmp_path / 'b.csv')
        with open(path, 'w') as f:
            f.write('\n\n')
        assert file_io.read_csv_any_encoding(path).shape == (0, 0)


class TestAtomicSave:
    def test_csv_save_keeps_permissions_and_no_temp_left(self, tmp_path):
        path = str(tmp_path / 'p.csv')
        file_io.save_csv(path, pd.DataFrame({'a': [1]}))
        os.chmod(path, 0o664)
        file_io.save_csv(path, pd.DataFrame({'a': [2]}))
        assert os.stat(path).st_mode & 0o777 == 0o664
        assert sorted(os.listdir(tmp_path)) == ['p.csv']

    def test_csv_save_failure_leaves_original_intact(self, tmp_path, monkeypatch):
        path = str(tmp_path / 'p.csv')
        file_io.save_csv(path, pd.DataFrame({'a': [1]}))
        before = open(path, 'rb').read()

        def boom(self, *a, **k):
            raise OSError('disk full')
        monkeypatch.setattr(pd.DataFrame, 'to_csv', boom)
        import pytest
        with pytest.raises(OSError):
            file_io.save_csv(path, pd.DataFrame({'a': [2]}))
        assert open(path, 'rb').read() == before
        assert sorted(os.listdir(tmp_path)) == ['p.csv']

    def test_xlsx_save_keeps_permissions(self, tmp_path):
        path = str(tmp_path / 'p.xlsx')
        file_io.save_workbook(path, {'S': pd.DataFrame({'a': [1]})})
        os.chmod(path, 0o664)
        file_io.save_workbook(path, {'S': pd.DataFrame({'a': [2]})})
        assert os.stat(path).st_mode & 0o777 == 0o664
        assert sorted(os.listdir(tmp_path)) == ['p.xlsx']


class TestSheetNameValidation:
    def test_rules(self):
        assert file_io.check_sheet_name('') is not None
        assert file_io.check_sheet_name('   ') is not None
        assert file_io.check_sheet_name('a' * 31) is None
        assert file_io.check_sheet_name('a' * 32) is not None
        for ch in '\\/?*[]:':
            assert file_io.check_sheet_name('a' + ch) is not None, ch
        assert file_io.check_sheet_name('Abc', ['abc']) is not None   # 不区分大小写
        assert file_io.check_sheet_name('ok', ['x']) is None

    def test_save_rejects_colliding_long_names_instead_of_overwriting(self, tmp_path):
        import pytest
        path = str(tmp_path / 'x.xlsx')
        long_a = 'S' * 31 + 'x'
        long_b = 'S' * 31 + 'y'
        with pytest.raises(ValueError):
            file_io.save_workbook(path, {long_a: pd.DataFrame({'a': [1]}),
                                         long_b: pd.DataFrame({'a': [2]})})
        assert not os.path.exists(path)

    def test_save_rejects_case_duplicate(self, tmp_path):
        import pytest
        with pytest.raises(ValueError):
            file_io.save_workbook(str(tmp_path / 'x.xlsx'),
                                  {'S': pd.DataFrame({'a': [1]}),
                                   's': pd.DataFrame({'a': [2]})})
