"""保存不再重建工作簿：透视表之类 openpyxl 能读回的东西必须原样保留。

用条件格式/数据验证/列宽/冻结窗格/合并单元格/图表当替身来验证——它们和
透视表走的是同一条"openpyxl 读回再写出"的路径。
"""
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import numpy as np
import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from PyQt6.QtWidgets import QApplication, QMessageBox

_app = QApplication.instance() or QApplication([])

from qtui import file_io
from qtui.main_window import MainWindow


def make_rich_xlsx(path, rows=5):
    wb = Workbook()
    ws = wb.active
    ws.title = "数据"
    ws.append(["a", "b"])
    for i in range(1, rows + 1):
        ws.append([i, i * 2])
    ws.column_dimensions["A"].width = 33
    ws.freeze_panes = "A2"
    ws.merge_cells("D1:E1")
    ws.conditional_formatting.add(
        "A2:A99", CellIsRule(operator="greaterThan", formula=["2"],
                             fill=PatternFill(bgColor="FFC7CE")))
    dv = DataValidation(type="list", formula1='"x,y"')
    ws.add_data_validation(dv)
    dv.add("C2:C99")
    chart = BarChart()
    chart.add_data(Reference(ws, min_col=1, min_row=1, max_row=rows + 1),
                   titles_from_data=True)
    ws.add_chart(chart, "G2")
    other = wb.create_sheet("原样")
    other.append(["k", "v"])
    other.append([1, "=A2*10"])
    other["B3"] = "别动我"
    wb.save(path)
    return path


def features(path, sheet="数据"):
    ws = load_workbook(path)[sheet]
    return {
        "cf": len(ws.conditional_formatting._cf_rules),
        "dv": len(ws.data_validations.dataValidation),
        "width": ws.column_dimensions["A"].width,
        "freeze": ws.freeze_panes,
        "merged": [str(r) for r in ws.merged_cells.ranges],
        "charts": len(ws._charts),
    }


class TestPatchWorkbook:
    def test_keeps_everything_pandas_rebuild_loses(self, tmp_path):
        src = make_rich_xlsx(str(tmp_path / "src.xlsx"))
        before = features(src)
        df = pd.DataFrame({"a": [1, 2, 3, 4, 5], "b": [9, 4, 6, 8, 10]})
        file_io.patch_workbook(src, src, {"数据": df}, ["数据", "原样"])
        assert features(src) == before
        ws = load_workbook(src)["数据"]
        assert ws["B2"].value == 9          # 改动写进去了

    def test_untouched_sheet_keeps_its_formula(self, tmp_path):
        src = make_rich_xlsx(str(tmp_path / "src.xlsx"))
        df = pd.DataFrame({"a": [1], "b": [2]})
        file_io.patch_workbook(src, src, {"数据": df}, ["数据", "原样"])
        ws = load_workbook(src)["原样"]
        assert ws["B2"].value == "=A2*10"
        assert ws["B3"].value == "别动我"

    def test_shrinking_table_deletes_leftover_rows(self, tmp_path):
        src = make_rich_xlsx(str(tmp_path / "src.xlsx"), rows=8)
        df = pd.DataFrame({"a": [1, 2], "b": [3, 4]})
        file_io.patch_workbook(src, src, {"数据": df}, ["数据", "原样"])
        ws = load_workbook(src)["数据"]
        assert ws.max_row == 3
        assert ws["A4"].value is None

    def test_header_formula_is_kept_not_flattened(self, tmp_path):
        src = str(tmp_path / "h.xlsx")
        wb = Workbook(); ws = wb.active; ws.title = "S"
        ws["A1"] = "=CONCATENATE(\"col\",\"A\")"
        ws["B1"] = "b"
        ws.append([1, 2])
        wb.save(src)
        df = pd.DataFrame({"colA": [7], "b": [8]})
        result = file_io.patch_workbook(src, src, {"S": df}, ["S"])
        ws = load_workbook(src)["S"]
        assert ws["A1"].value == "=CONCATENATE(\"col\",\"A\")"
        assert ws["B1"].value == "b"
        assert result["kept_header_formulas"] == 1

    def test_formulas_and_colors_written(self, tmp_path):
        src = make_rich_xlsx(str(tmp_path / "src.xlsx"))
        df = pd.DataFrame({"a": [1, 2], "b": [3, 4]})
        file_io.patch_workbook(src, src, {"数据": df}, ["数据", "原样"],
                               formulas={"数据": {(0, 1): "=A2*3"}},
                               cell_colors={"数据": {(-1, 0): "#ff0000",
                                                     (1, 0): "#00ff00"}})
        ws = load_workbook(src)["数据"]
        assert ws["B2"].value == "=A2*3"
        assert ws["A1"].fill.start_color.rgb == "FFFF0000"
        assert ws["A3"].fill.start_color.rgb == "FF00FF00"

    def test_cleared_color_is_removed(self, tmp_path):
        src = str(tmp_path / "c.xlsx")
        wb = Workbook(); ws = wb.active; ws.title = "S"
        ws.append(["a"]); ws.append([1])
        ws["A2"].fill = PatternFill(start_color="FF00FF00", end_color="FF00FF00",
                                    fill_type="solid")
        wb.save(src)
        df = pd.DataFrame({"a": [1]})
        file_io.patch_workbook(src, src, {"S": df}, ["S"], cell_colors={"S": {}})
        assert load_workbook(src)["S"]["A2"].fill.fill_type is None

    def test_sheet_add_delete_and_order(self, tmp_path):
        src = make_rich_xlsx(str(tmp_path / "src.xlsx"))
        df = pd.DataFrame({"x": [1]})
        file_io.patch_workbook(src, src, {"新表": df}, ["新表", "数据"])
        wb = load_workbook(src)
        assert wb.sheetnames == ["新表", "数据"]   # 原样 sheet 被删除，顺序对齐

    def test_numpy_and_missing_values(self, tmp_path):
        src = str(tmp_path / "n.xlsx")
        wb = Workbook(); ws = wb.active; ws.title = "S"
        ws.append(["a", "b", "c"]); ws.append([0, 0, 0])
        wb.save(src)
        df = pd.DataFrame({"a": [np.int64(5)], "b": [np.nan],
                           "c": [pd.Timestamp("2026-01-02 03:04:05")]})
        file_io.patch_workbook(src, src, {"S": df}, ["S"])
        ws = load_workbook(src)["S"]
        assert ws["A2"].value == 5 and ws["B2"].value is None
        assert str(ws["C2"].value) == "2026-01-02 03:04:05"

    def test_save_as_copy_keeps_source_intact(self, tmp_path):
        src = make_rich_xlsx(str(tmp_path / "src.xlsx"))
        dest = str(tmp_path / "copy.xlsx")
        df = pd.DataFrame({"a": [7], "b": [8]})
        file_io.patch_workbook(src, dest, {"数据": df}, ["数据", "原样"])
        assert features(dest) == features(src)          # 副本继承全部设置
        assert load_workbook(src)["数据"]["A2"].value == 1   # 原文件没被动


class TestLossyParts:
    def test_plain_file_has_nothing_to_warn_about(self, tmp_path):
        src = make_rich_xlsx(str(tmp_path / "src.xlsx"))
        assert file_io.xlsx_lossy_parts(src) == []

    def test_macro_workbook_is_flagged(self, tmp_path):
        import zipfile
        src = make_rich_xlsx(str(tmp_path / "src.xlsx"))
        with zipfile.ZipFile(src, "a") as z:
            z.writestr("xl/vbaProject.bin", b"\x00")
        assert "宏（VBA）" in file_io.xlsx_lossy_parts(src)


class TestSaveThroughWindow:
    def test_editing_a_cell_keeps_workbook_features(self, tmp_path, monkeypatch):
        src = make_rich_xlsx(str(tmp_path / "src.xlsx"))
        before = features(src)
        win = MainWindow()
        win.load_file(src)
        _app.processEvents()
        win.model.setData(win.model.index(1, 1), "99")
        assert win.save_file() is True
        _app.processEvents()
        assert features(src) == before
        ws = load_workbook(src)["数据"]
        assert ws["B2"].value == 99
        assert ws["A2"].value == 1
        win.model.modified = False
        win.close()

    def test_lossy_features_ask_before_saving(self, tmp_path, monkeypatch):
        import zipfile
        src = make_rich_xlsx(str(tmp_path / "src.xlsx"))
        with zipfile.ZipFile(src, "a") as z:
            z.writestr("xl/slicers/slicer1.xml", b"<x/>")
        win = MainWindow()
        win.load_file(src)
        _app.processEvents()
        win.model.setData(win.model.index(1, 1), "99")
        asked = []
        monkeypatch.setattr(QMessageBox, "warning",
                            lambda *a, **k: (asked.append(a[2]),
                                             QMessageBox.StandardButton.Cancel)[1])
        assert win.save_file() is False
        assert asked and "切片器" in asked[0]
        assert load_workbook(src)["数据"]["B2"].value == 2   # 取消后原文件没动
        win.model.modified = False
        win.close()
